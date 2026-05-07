#!/usr/bin/env python3
"""
BLT Trading price scraper: variant-first HTML generator.

Downloads the public BLT Google Sheet, parses per-category tabs, and generates
per-category HTML files with GPT-optimized chunking: one <section> per device
variant (model + storage + lock), with all grade prices inside.
"""
import os
import re
import sys
import datetime
from pathlib import Path
from collections import defaultdict
import requests
from openpyxl import load_workbook

# Constants
XLSX_URL = "https://docs.google.com/spreadsheets/d/1Cg1ZuaILPJbEdqbgd01PjUESbZh3h3TaK-NB9oSBMLc/export?format=xlsx"
BLT_XLSX_PATH = "/tmp/blt.xlsx"
OUTDIR = Path(os.environ.get("BLT_OUTDIR", "."))
OUTDIR.mkdir(parents=True, exist_ok=True)

# ===== Data structures =====

class PriceEntry:
    """One price entry: device variant + condition + price."""
    def __init__(self, category, model, storage, lock, condition, price, new_used="USED"):
        self.category = category
        self.model = model
        self.storage = storage
        self.lock = lock
        self.condition = condition
        self.price = price
        self.new_used = new_used
        self.variant_key = (self.model, self.storage, self.lock)

    def identifier(self):
        """Human-readable identifier."""
        parts = [self.model]
        if self.storage:
            parts.append(self.storage)
        parts.append(self.lock)
        return " ".join(parts)

    def __repr__(self):
        return f"PriceEntry({self.category}, {self.identifier()}, {self.condition}, ${self.price})"


# ===== Parsing functions =====

def download_xlsx():
    """Download XLSX from Google Sheet."""
    print(f"Downloading {XLSX_URL}...")
    r = requests.get(XLSX_URL)
    r.raise_for_status()
    Path(BLT_XLSX_PATH).write_bytes(r.content)
    print(f"Saved to {BLT_XLSX_PATH}")


def parse_iphone_used(ws):
    """Parse 'Used iphone ' tab (USED iPhones)."""
    entries = []
    max_row = ws.max_row

    for row_idx in range(2, max_row + 1):
        cell_a = ws[f"A{row_idx}"].value
        if not cell_a or not isinstance(cell_a, str):
            continue

        cell_a = cell_a.strip()
        if not cell_a.startswith("iPhone "):
            continue

        m = re.match(r"iPhone\s+(.+?)\s+(\d+(?:GB|TB))\s+(Unlocked|Locked)$", cell_a)
        if not m:
            continue

        model = f"iPhone {m.group(1)}"
        storage = m.group(2)
        lock_str = m.group(3)

        conditions = [
            ("SWAP HSO", "B"),
            ("Grade A", "C"),
            ("Grade B", "D"),
            ("Grade C", "E"),
            ("Grade D", "F"),
            ("DOA", "G"),
        ]

        for cond_name, col in conditions:
            price_cell = ws[f"{col}{row_idx}"].value
            if price_cell is None or price_cell == "-":
                continue
            try:
                price = int(float(str(price_cell).replace(",", "")))
            except (ValueError, TypeError):
                continue

            lock = "Carrier Locked" if lock_str == "Locked" else "Unlocked"

            entries.append(PriceEntry(
                category="iphone-used",
                model=model,
                storage=storage,
                lock=lock,
                condition=cond_name,
                price=price,
                new_used="USED"
            ))

    return entries


def parse_iphone_new(ws):
    """Parse 'New Iphone' tab (NEW iPhones)."""
    entries = []
    max_row = ws.max_row

    for row_idx in range(2, max_row + 1):
        cell_a = ws[f"A{row_idx}"].value
        if not cell_a or not isinstance(cell_a, str):
            continue

        cell_a = cell_a.strip()
        if not cell_a.startswith("iPhone "):
            continue

        m = re.match(r"iPhone\s+(.+?)\s+(\d+(?:GB|TB))\s+(Unlocked|Locked)$", cell_a)
        if not m:
            continue

        model = f"iPhone {m.group(1)}"
        storage = m.group(2)
        lock_str = m.group(3)
        lock = "Carrier Locked" if lock_str == "Locked" else "Unlocked"

        sealed_cells = [
            ws[f"B{row_idx}"].value,
            ws[f"C{row_idx}"].value,
            ws[f"D{row_idx}"].value,
        ]
        open_cell = ws[f"E{row_idx}"].value
        activated_cell = ws[f"F{row_idx}"].value

        sealed_price = None
        for cell in sealed_cells:
            if cell and cell != "-":
                try:
                    sealed_price = int(float(str(cell).replace(",", "")))
                    break
                except (ValueError, TypeError):
                    pass

        if sealed_price:
            entries.append(PriceEntry(
                category="iphone-new",
                model=model,
                storage=storage,
                lock=lock,
                condition="Sealed",
                price=sealed_price,
                new_used="NEW"
            ))

        if open_cell and open_cell != "-":
            try:
                open_price = int(float(str(open_cell).replace(",", "")))
                entries.append(PriceEntry(
                    category="iphone-new",
                    model=model,
                    storage=storage,
                    lock=lock,
                    condition="Open Box",
                    price=open_price,
                    new_used="NEW"
                ))
            except (ValueError, TypeError):
                pass

        if activated_cell and activated_cell != "-":
            try:
                activated_price = int(float(str(activated_cell).replace(",", "")))
                entries.append(PriceEntry(
                    category="iphone-new",
                    model=model,
                    storage=storage,
                    lock=lock,
                    condition="Sealed (Activated)",
                    price=activated_price,
                    new_used="NEW"
                ))
            except (ValueError, TypeError):
                pass

    return entries


def parse_ipad_used(ws):
    """Parse 'Used ipads ' tab (USED iPads)."""
    entries = []
    max_row = ws.max_row

    for row_idx in range(2, max_row + 1):
        cell_a = ws[f"A{row_idx}"].value
        if not cell_a:
            continue

        cell_a = str(cell_a).strip()

        if not cell_a.startswith("iPad"):
            continue

        conditions = [
            ("Grade A", "B"),
            ("Grade B+", "C"),
            ("Grade B", "D"),
            ("Grade C", "E"),
            ("Grade D", "F"),
            ("DOA", "G"),
        ]

        storage = None
        lock = "WiFi"

        m = re.search(r"(\d+(?:GB|TB))", cell_a)
        if m:
            storage = m.group(1)

        if "Cellular" in cell_a:
            lock = "Cellular"
        elif "Verizon" in cell_a:
            lock = "Verizon"

        model_match = re.match(r"(iPad[^,\d]*)", cell_a)
        model = model_match.group(1).strip() if model_match else "iPad"

        for cond_name, col in conditions:
            price_cell = ws[f"{col}{row_idx}"].value
            if price_cell is None or price_cell == "-":
                continue
            try:
                price = int(float(str(price_cell).replace(",", "")))
            except (ValueError, TypeError):
                continue

            entries.append(PriceEntry(
                category="ipad",
                model=model,
                storage=storage,
                lock=lock,
                condition=cond_name,
                price=price,
                new_used="USED"
            ))

    return entries


def parse_ipad_new(ws):
    """Parse 'New ipads' tab (NEW iPads)."""
    entries = []
    max_row = ws.max_row

    for row_idx in range(2, max_row + 1):
        cell_a = ws[f"A{row_idx}"].value
        if not cell_a:
            continue

        cell_a = str(cell_a).strip()
        if not cell_a.startswith("iPad"):
            continue

        storage = None
        lock = "WiFi"

        m = re.search(r"(\d+(?:GB|TB))", cell_a)
        if m:
            storage = m.group(1)

        if "Cellular" in cell_a:
            lock = "Cellular"
        elif "Verizon" in cell_a:
            lock = "Verizon"

        model_match = re.match(r"(iPad[^,\d]*)", cell_a)
        model = model_match.group(1).strip() if model_match else "iPad"

        conditions = [
            ("Sealed", "B"),
            ("Open Box", "C"),
            ("Sealed (Activated)", "D"),
        ]

        for cond_name, col in conditions:
            price_cell = ws[f"{col}{row_idx}"].value
            if price_cell is None or price_cell == "-":
                continue
            try:
                price = int(float(str(price_cell).replace(",", "")))
            except (ValueError, TypeError):
                continue

            entries.append(PriceEntry(
                category="ipad",
                model=model,
                storage=storage,
                lock=lock,
                condition=cond_name,
                price=price,
                new_used="NEW"
            ))

    return entries


def parse_samsung(ws):
    """Parse 'Samsung' tab."""
    entries = []
    max_row = ws.max_row
    current_model = None

    for row_idx in range(2, max_row + 1):
        cell_a = ws[f"A{row_idx}"].value
        cell_b = ws[f"B{row_idx}"].value

        if not cell_a and not cell_b:
            continue

        cell_a = str(cell_a).strip() if cell_a else ""
        cell_b = str(cell_b).strip() if cell_b else ""

        if cell_a and "Unlocked" in cell_b:
            current_model = cell_a
            lock = "Unlocked"
        elif not cell_a and "Locked" in cell_b:
            lock = "Carrier Locked"
        else:
            continue

        if not current_model:
            continue

        conditions = [
            ("Sealed", "C", "NEW"),
            ("Grade A", "D", "USED"),
            ("Grade B", "E", "USED"),
            ("Grade C", "F", "USED"),
            ("Grade D", "G", "USED"),
            ("DOA", "H", "USED"),
        ]

        for cond_name, col, new_used in conditions:
            price_cell = ws[f"{col}{row_idx}"].value
            if price_cell is None or price_cell == "-":
                continue
            try:
                price = int(float(str(price_cell).replace(",", "")))
            except (ValueError, TypeError):
                continue

            entries.append(PriceEntry(
                category="samsung",
                model=current_model,
                storage=None,
                lock=lock,
                condition=cond_name,
                price=price,
                new_used=new_used
            ))

    return entries


def parse_watch(ws):
    """Parse 'Apple watch ' tab."""
    entries = []
    max_row = ws.max_row
    current_model = None

    for row_idx in range(2, max_row + 1):
        cell_a = ws[f"A{row_idx}"].value
        if not cell_a:
            continue

        cell_a = str(cell_a).strip()

        if re.match(r"^Series \d+$", cell_a):
            current_model = cell_a
            continue

        if not cell_a.startswith("Series") and current_model:
            variant = cell_a
        else:
            continue

        conditions = [
            ("Sealed", "B", "NEW"),
            ("Open Box", "C", "NEW"),
            ("Grade A", "D", "USED"),
            ("Grade B", "E", "USED"),
            ("Grade C", "F", "USED"),
            ("Grade D", "G", "USED"),
        ]

        for cond_name, col, new_used in conditions:
            price_cell = ws[f"{col}{row_idx}"].value
            if price_cell is None or price_cell == "-":
                continue
            try:
                price = int(float(str(price_cell).replace(",", "")))
            except (ValueError, TypeError):
                continue

            entries.append(PriceEntry(
                category="watch",
                model=variant,
                storage=None,
                lock=None,
                condition=cond_name,
                price=price,
                new_used=new_used
            ))

    return entries


def parse_gaming(ws):
    """Parse 'switch- PS5' tab."""
    entries = []
    max_row = ws.max_row

    for row_idx in range(2, max_row + 1):
        cell_a = ws[f"A{row_idx}"].value
        if not cell_a:
            continue

        cell_a = str(cell_a).strip()
        cell_a = cell_a.lstrip("•").strip()

        if not cell_a:
            continue

        conditions = [
            ("Sealed", "B", "NEW"),
            ("Open Box", "C", "NEW"),
        ]

        for cond_name, col, new_used in conditions:
            price_cell = ws[f"{col}{row_idx}"].value
            if price_cell is None or price_cell == "-":
                continue
            try:
                price = int(float(str(price_cell).replace(",", "")))
            except (ValueError, TypeError):
                continue

            entries.append(PriceEntry(
                category="gaming",
                model=cell_a,
                storage=None,
                lock=None,
                condition=cond_name,
                price=price,
                new_used=new_used
            ))

    return entries


# ===== HTML generation =====

def _disambig_for(model):
    """Build a 'NOT X / NOT Y' disambiguation string for a model."""
    m = re.match(r"^iPhone (\d+)\s*(.*)$", model)
    if m:
        num = int(m.group(1))
        suffix = m.group(2).strip()
        siblings = []
        if suffix == "":
            siblings = [f"iPhone {num} Pro", f"iPhone {num} Pro Max",
                        f"iPhone {num} Plus", f"iPhone {num}E", f"iPhone {num} Air"]
        elif suffix.lower() == "pro":
            siblings = [f"iPhone {num} Pro Max", f"iPhone {num} (base)", f"iPhone {num} Plus"]
        elif suffix.lower() == "pro max":
            siblings = [f"iPhone {num} Pro", f"iPhone {num} (base)", f"iPhone {num} Plus"]
        elif suffix.lower() == "plus":
            siblings = [f"iPhone {num} (base)", f"iPhone {num} Pro", f"iPhone {num} Pro Max"]
        elif suffix.upper() in ("E",):
            siblings = [f"iPhone {num} (base)", f"iPhone {num} Pro"]
        elif suffix.upper() == "AIR":
            siblings = [f"iPhone {num} Pro", f"iPhone {num} (base)"]
        siblings.append(f"iPhone {num-1} {suffix}".strip())
        siblings.append(f"iPhone {num+1} {suffix}".strip())
        siblings = [s for s in siblings if s != model]
        return " — ".join(["DO NOT confuse with " + s for s in siblings[:5]])
    return ""


def _short_lock(lock):
    if not lock:
        return ""
    if "Carrier" in lock:
        return "Carrier Locked"
    if lock == "Unlocked":
        return "SIM Unlocked"
    return lock


def _full_lock_label(lock):
    if not lock:
        return ""
    if "Carrier" in lock:
        return "AT&amp;T / T-Mobile / Sprint / Verizon / US Cellular Carrier Locked"
    if lock == "Unlocked":
        return "SIM Unlocked / Factory Unlocked"
    return lock


GRADE_DESC = {
    "Grade A": "no scratches, mint, like new",
    "Grade B": "light use, no cracks, fully functional",
    "Grade B+": "light use, no cracks, fully functional",
    "Grade C": "cracked screen, hairline crack, heavy scratches",
    "Grade D": "bad LCD, dead pixels, lines on screen",
    "DOA": "won't power, water damage",
    "Sealed": "factory-sealed, brand new, never opened",
    "Open Box": "opened, brand new, never used",
    "Sealed (Activated)": "sealed in box, activation started",
    "SWAP HSO": "factory-fresh, never used, no original box — requires seller to explicitly say all three: brand new + never used + no box",
}

GRADE_ORDER = {
    "Grade A": 0, "Grade B": 1, "Grade B+": 2, "Grade C": 3, "Grade D": 4,
    "DOA": 5, "SWAP HSO": 6, "Sealed": 0, "Open Box": 1, "Sealed (Activated)": 2,
}


def render_section(variant_key, entries):
    """Compact section: ~60 words. Fits 3-4 per GHL chunk so the bot can hold
    multiple variants AND the defaults block in a single retrieval payload."""
    model, storage, lock = variant_key

    short_lock = _short_lock(lock)
    full_lock = _full_lock_label(lock)

    short_id_parts = [model]
    if storage:
        short_id_parts.append(storage)
    if short_lock:
        short_id_parts.append(short_lock)
    short_id = " ".join(short_id_parts)

    header_parts = [model]
    if storage:
        header_parts.append(storage)
    if full_lock:
        header_parts.append(full_lock)
    header_id = " ".join(header_parts)

    entries_sorted = sorted(entries, key=lambda e: GRADE_ORDER.get(e.condition, 99))
    default_entry = next(
        (e for e in entries_sorted if e.condition in ("Grade A", "Sealed")),
        entries_sorted[0] if entries_sorted else None,
    )
    default_price = f"${default_entry.price}" if default_entry else "$?"
    default_cond = default_entry.condition if default_entry else "Grade A"

    # Compact "NOT X" disambiguation: just the 2-3 closest siblings inline.
    siblings = []
    m = re.match(r"^iPhone (\d+)\s*(.*)$", model)
    if m:
        num = int(m.group(1))
        suffix = m.group(2).strip()
        if suffix.lower() == "pro":
            siblings = [f"iPhone {num} Pro Max", f"iPhone {num} (base)"]
        elif suffix.lower() == "pro max":
            siblings = [f"iPhone {num} Pro", f"iPhone {num-1} Pro Max"]
        elif suffix.lower() == "plus":
            siblings = [f"iPhone {num} (base)", f"iPhone {num} Pro"]
        elif suffix == "":
            siblings = [f"iPhone {num} Pro", f"iPhone {num} Plus"]
        elif suffix.upper() == "E":
            siblings = [f"iPhone {num} (base)", f"iPhone {num} Pro"]
    not_phrase = ""
    if siblings:
        not_phrase = " NOT " + " or ".join(siblings) + "."

    html = []
    html.append("<section>")
    html.append(f"<h2>{header_id}</h2>")
    html.append(
        f"<p><strong>{short_id}</strong> default <strong>{default_price}</strong> ({default_cond}).{not_phrase}</p>"
    )

    html.append("<ul>")
    short_grade_desc = {
        "Grade A": "no scratches",
        "Grade B": "light use",
        "Grade B+": "light use",
        "Grade C": "cracked / heavy scratches",
        "Grade D": "bad LCD",
        "DOA": "won't power",
        "Sealed": "factory-sealed",
        "Open Box": "opened, never used",
        "Sealed (Activated)": "sealed, activated",
        "SWAP HSO": "factory-fresh no box (req. 'brand new + never used + no box')",
    }
    for entry in entries_sorted:
        cond = entry.condition
        price = f"${entry.price}"
        desc = short_grade_desc.get(cond, "")
        marker = " [DEFAULT]" if cond == default_cond else ""
        html.append(f"<li>{cond}: {price} ({desc}){marker}</li>")
    html.append("</ul>")
    html.append("</section>")
    return "\n".join(html)


def render_per_model_section(model, entries):
    """Compact per-model section: one <section> per iPhone model, all variants
    and conditions inside. ~80-120 words. GHL chunks will then group multiple
    complete models per chunk instead of slicing one model across chunks.
    """
    # Disambiguation siblings
    siblings = []
    m = re.match(r"^iPhone (\d+)\s*(.*)$", model)
    if m:
        num = int(m.group(1))
        suffix = m.group(2).strip()
        if suffix.lower() == "pro":
            siblings = [f"iPhone {num} Pro Max", f"iPhone {num} (base)"]
        elif suffix.lower() == "pro max":
            siblings = [f"iPhone {num} Pro", f"iPhone {num-1} Pro Max"]
        elif suffix.lower() == "plus":
            siblings = [f"iPhone {num} (base)", f"iPhone {num} Pro"]
        elif suffix == "":
            siblings = [f"iPhone {num} Pro", f"iPhone {num} Plus"]
        elif suffix.upper() == "E":
            siblings = [f"iPhone {num} (base)", f"iPhone {num} Pro"]
    not_phrase = ""
    if siblings:
        not_phrase = " NOT " + " or ".join(siblings) + "."

    # Group entries by (storage, lock)
    by_variant = defaultdict(list)
    for e in entries:
        by_variant[(e.storage, e.lock)].append(e)

    # Find smallest-storage Grade A defaults for unlocked and carrier-locked
    def grade_a(lock):
        candidates = [
            (s, lst) for (s, l), lst in by_variant.items()
            if l == lock and any(x.condition == "Grade A" for x in lst)
        ]
        if not candidates: return None
        candidates.sort(key=lambda x: _storage_kb(x[0]))
        s, lst = candidates[0]
        a = next(x for x in lst if x.condition == "Grade A")
        return (s, a.price)

    unlock_default = grade_a("Unlocked")
    locked_default = grade_a("Carrier Locked")

    # Find smallest-storage Sealed for carrier-locked (NEW Sealed default)
    sealed_default = None
    sealed_candidates = [
        (s, lst) for (s, l), lst in by_variant.items()
        if l == "Carrier Locked" and any(x.condition == "Sealed" for x in lst)
    ]
    if sealed_candidates:
        sealed_candidates.sort(key=lambda x: _storage_kb(x[0]))
        s, lst = sealed_candidates[0]
        sealed = next(x for x in lst if x.condition == "Sealed")
        sealed_default = (s, sealed.price)

    html = []
    html.append("<section>")
    html.append(f"<h2>{model} — All Variants and Conditions</h2>")

    # DEFAULTS callout (always first, model name repeated for chunk-fragment safety)
    default_parts = []
    if unlock_default:
        s, p = unlock_default
        default_parts.append(f"SIM Unlocked Grade A {s} ${p}")
    if locked_default:
        s, p = locked_default
        default_parts.append(f"Carrier-Locked Grade A {s} ${p}")
    if sealed_default:
        s, p = sealed_default
        default_parts.append(f"NEW Sealed Carrier-Locked {s} ${p}")
    if default_parts:
        html.append(
            f"<p><strong>{model} DEFAULTS:</strong> " + "; ".join(default_parts) + f".{not_phrase}</p>"
        )

    # Variant detail lines — one per (storage, lock) combo, all conditions inline
    grade_order_local = {
        "Grade A": 0, "Grade B": 1, "Grade B+": 2, "Sealed": 3, "Open Box": 4,
        "Sealed (Activated)": 5, "Grade C": 6, "Grade D": 7, "DOA": 8, "SWAP HSO": 9,
    }
    short_cond = {
        "Grade A": "A", "Grade B": "B", "Grade B+": "B+", "Grade C": "C",
        "Grade D": "D", "DOA": "DOA", "SWAP HSO": "HSO",
        "Sealed": "Sealed", "Open Box": "OpenBox", "Sealed (Activated)": "SealedActivated",
    }

    # Sort variants: storage ascending, then Unlocked before Carrier Locked
    variant_keys = sorted(
        by_variant.keys(),
        key=lambda x: (_storage_kb(x[0]), 0 if x[1] == "Unlocked" else 1)
    )

    for (storage, lock) in variant_keys:
        ents = sorted(by_variant[(storage, lock)], key=lambda e: grade_order_local.get(e.condition, 99))
        bits = []
        for e in ents:
            label = short_cond.get(e.condition, e.condition)
            marker = " [DEFAULT]" if e.condition == "Grade A" else ""
            bits.append(f"{label} ${e.price}{marker}")
        lock_label = "SIM Unlocked" if lock == "Unlocked" else "Carrier-Locked (AT&T/T-Mobile/Sprint/Verizon/US Cellular)"
        html.append(
            f"<p><strong>{model} {storage} {lock_label}:</strong> {', '.join(bits)}.</p>"
        )

    html.append("</section>")
    return "\n".join(html)



def _storage_kb(s):
    if not s:
        return 99999
    m = re.match(r"(\d+)\s*(GB|TB)", s)
    if not m:
        return 99999
    n = int(m.group(1))
    return n * (1024 if m.group(2) == "TB" else 1)


def render_quick_answers(all_entries):
    """The Quick Answers block — one canonical default sentence per model.
    Goes at the very top of prices.html so short queries like "16 pro" or "17"
    or "16 pro max" land here first instead of inside a buried variant section."""
    html = []
    html.append('<h1>QUICK ANSWERS — DEFAULT BUYING PRICES (read this first)</h1>')
    html.append(
        "<p>This block is the canonical default for every device. When a seller types JUST a model name "
        "(e.g. \"17 pro\", \"16 pro max\", \"iPhone 13\", \"iPad Pro 11\", \"Series 9\", \"AirPods Pro 2\") "
        "with no storage, lock, or condition, quote the matching line below and only the matching line. "
        "The variant sections farther down on this page are ONLY for refinement when the seller adds more detail "
        "(different storage, carrier-locked, sealed, scratched, cracked, etc.).</p>"
    )
    html.append(
        "<p><strong>Rule:</strong> never reply \"that model isn't on our buying list\" for any iPhone 14, 15, 16, or 17 "
        "(any variant — base, Plus, Pro, Pro Max, Air, E). They are all listed below. If the seller's exact wording isn't found, "
        "ask them to repeat the full model name; do not escalate.</p>"
    )

    # ---------- iPhones ----------
    used_iphones = {}
    for e in all_entries:
        if e.category != "iphone-used":
            continue
        if e.condition != "Grade A" or e.lock != "Unlocked":
            continue
        used_iphones.setdefault(e.model, []).append(e)

    if used_iphones:
        html.append("<h2>iPhone (USED Grade A, smallest storage, SIM Unlocked) — naked-query defaults</h2>")
        # Order: 17 series, 16 series, 15 series, 14 series, 13 series, 12 series, etc.
        def _iphone_key(model):
            m = re.match(r"iPhone (\d+)\s*(.*)$", model)
            if not m:
                return (0, model)
            num = int(m.group(1))
            suffix = m.group(2).strip().lower()
            suffix_order = {"pro max": 0, "pro": 1, "air": 2, "plus": 3, "": 4, "e": 5, "mini": 6}
            return (-num, suffix_order.get(suffix, 99), model)
        for model in sorted(used_iphones.keys(), key=_iphone_key):
            entries_m = sorted(used_iphones[model], key=lambda e: _storage_kb(e.storage))
            best = entries_m[0]
            disambig = _disambig_for(model)
            line = (
                f"<p><strong>{model}</strong> — naked-query default <strong>${best.price}</strong> "
                f"({best.storage} SIM Unlocked Grade A). "
                f"Reply template: \"If it's SIM unlocked with no scratches at all, we can offer up to ${best.price} "
                f"for the {best.storage} {model}. Prices can change at any time at our discretion.\""
            )
            if disambig:
                line += f" <em>{disambig}.</em>"
            line += "</p>"
            html.append(line)

    # ---------- iPads ----------
    ipads = {}
    for e in all_entries:
        if e.category != "ipad":
            continue
        if e.condition != "Grade A":
            continue
        # iPads: prefer WiFi for naked default
        if e.lock not in ("WiFi", "Cellular"):
            continue
        ipads.setdefault(e.model, []).append(e)

    if ipads:
        html.append("<h2>iPad (USED Grade A, smallest storage, WiFi or Cellular) — naked-query defaults</h2>")
        for model in sorted(ipads.keys()):
            entries_m = sorted(ipads[model], key=lambda e: _storage_kb(e.storage))
            best = entries_m[0]
            html.append(
                f"<p><strong>{model}</strong> — naked-query default <strong>${best.price}</strong> "
                f"({best.storage} {best.lock} Grade A). "
                f"Reply template: \"With no scratches at all, we can offer up to ${best.price} for the "
                f"{best.storage} {model}. Prices can change at any time at our discretion.\"</p>"
            )

    # ---------- Samsung ----------
    samsung = {}
    for e in all_entries:
        if e.category != "samsung":
            continue
        if e.condition != "Grade A" or e.lock != "Unlocked":
            continue
        samsung.setdefault(e.model, []).append(e)

    if samsung:
        html.append("<h2>Samsung (USED Grade A, SIM Unlocked) — naked-query defaults</h2>")
        for model in sorted(samsung.keys()):
            entries_m = samsung[model]
            best = entries_m[0]
            html.append(
                f"<p><strong>{model}</strong> — naked-query default <strong>${best.price}</strong> "
                f"(SIM Unlocked Grade A). "
                f"Reply template: \"If it's SIM unlocked with no scratches at all, we can offer up to ${best.price} "
                f"for the {model}. Prices can change at any time at our discretion.\"</p>"
            )

    # ---------- Apple Watch ----------
    watch = {}
    for e in all_entries:
        if e.category != "watch":
            continue
        if e.condition != "Grade A":
            continue
        watch.setdefault(e.model, []).append(e)

    if watch:
        html.append("<h2>Apple Watch (USED Grade A) — naked-query defaults</h2>")
        for model in sorted(watch.keys()):
            best = watch[model][0]
            html.append(
                f"<p><strong>Apple Watch {model}</strong> — naked-query default <strong>${best.price}</strong> "
                f"(Grade A). Reply template: \"With no scratches at all, we can offer up to ${best.price} for the "
                f"Apple Watch {model}. Prices can change at any time at our discretion.\"</p>"
            )

    # ---------- Gaming ----------
    gaming = {}
    for e in all_entries:
        if e.category != "gaming":
            continue
        if e.condition != "Sealed":
            continue
        gaming.setdefault(e.model, []).append(e)

    if gaming:
        html.append("<h2>Gaming consoles (NEW Sealed) — naked-query defaults</h2>")
        for model in sorted(gaming.keys()):
            best = gaming[model][0]
            html.append(
                f"<p><strong>{model}</strong> — naked-query default <strong>${best.price}</strong> (Sealed). "
                f"Reply template: \"If it's sealed in box, we can offer up to ${best.price} for the {model}.\"</p>"
            )

    return "\n".join(html)


def render_category_html(category, title, entries):
    """Render a per-category HTML file (iphone-used.html, ipad.html, etc.)."""
    today_utc = datetime.datetime.now(datetime.timezone.utc).isoformat() + "Z"

    # iphone-new.html is now a stub — all NEW prices have been merged into
    # iphone-used.html sections. We keep this file so any cached crawler URL
    # still resolves, but it tells the bot to look elsewhere.
    if category == "iphone-new":
        return (
            "<!DOCTYPE html>\n"
            "<html lang=\"en\"><head><meta charset=\"UTF-8\">\n"
            "<title>BLT Trading — iPhone Buying Prices (NEW)</title></head>\n"
            "<body>\n"
            "<h1>NEW iPhone prices have moved</h1>\n"
            f"<p><strong>Last Updated:</strong> {today_utc}</p>\n"
            "<p>NEW Sealed, Open Box, and Sealed (Activated) prices are now listed inside the matching variant section in iphone-used.html — one chunk per variant contains ALL conditions (Grade A DEFAULT, NEW Sealed, Open Box, etc.). Quoting from this file directly is no longer accurate.</p>\n"
            "<p>If a seller says \"sealed\", \"brand new\", \"still in box\", or \"never opened\" — quote the NEW Sealed line inside the matching variant section in iphone-used.html.</p>\n"
            "<p>If a seller says \"open box\" — quote the NEW Open Box line inside the matching variant section in iphone-used.html.</p>\n"
            "<p>Otherwise — quote the USED Grade A line, which is the DEFAULT for every variant.</p>\n"
            "</body></html>\n"
        )

    html = []
    html.append("<!DOCTYPE html>")
    html.append('<html lang="en"><head><meta charset="UTF-8">')
    html.append(f"<title>{title}</title></head>")
    html.append("<body>")
    html.append(f"<h1>{title}</h1>")
    html.append(f"<p><strong>Last Updated:</strong> {today_utc}</p>")
    html.append("")

    by_variant = defaultdict(list)
    for entry in entries:
        by_variant[entry.variant_key].append(entry)

    variant_order = sorted(by_variant.keys())

    # iPad-specific clarifying-question flow. Unlike iPhones (where smallest-storage
    # SIM-Unlocked Grade A is a safe default), iPads vary too much in price across
    # storage and WiFi/Cellular for blind quoting to be useful.
    if category == "ipad":
        html.append("<h2>iPad clarifying-question rule (read this BEFORE quoting any iPad)</h2>")
        html.append("<p><strong>For iPad queries, do NOT quote a price upfront.</strong> iPad price varies $200+ by model + storage + WiFi/Cellular, so a blind quote is misleading. Instead, ask ONE combined question first, THEN quote from the matching variant section below.</p>")
        html.append("<p><strong>Reply for any naked iPad query</strong> (e.g. \"ipad\", \"ipad pro\", \"ipad air\", \"ipad mini\", \"ipad 9th gen\") <strong>→ ask:</strong></p>")
        html.append("<p>\"To get you a price, which iPad model and storage (e.g., iPad Pro 11\" M4 256GB or iPad Air 11\" M2 128GB), and is it WiFi or Cellular?\"</p>")
        html.append("<p><strong>Default assumption when seller specifies model + storage but NOT lock:</strong> Cellular. Quote the Cellular price and add \"except for Verizon\" if Verizon-locked Cellular has a different price.</p>")
        html.append("<p><strong>Default assumption when seller says \"sealed\" or \"brand new\":</strong> Cellular + \"except for Verizon\".</p>")
        html.append("<p><strong>USED iPad with no scratches:</strong> quote the Grade A price (single price, WiFi and Cellular often priced the same — see variant section).</p>")
        html.append("<p><strong>End every USED iPad quote with:</strong> \"Prices can change at any time at our discretion.\"</p>")
        html.append("")

        # Build a Query->Reply map for common naked iPad queries — each maps to the
        # clarifying question, not a price.
        ipad_query_aliases = [
            ("ipad", "Ambiguous \"ipad\" query. Ask which model and storage."),
            ("ipad pro", "Multiple iPad Pro generations and sizes exist."),
            ("ipad air", "Multiple iPad Air generations exist."),
            ("ipad mini", "Multiple iPad Mini generations exist."),
            ("ipad 9", "iPad 9th gen exists; confirm storage and lock."),
            ("ipad 10", "iPad 10th gen exists; confirm storage and lock."),
            ("ipad 11", "Could mean iPad 11th gen, iPad Pro 11\", or iPad Air 11\"."),
            ("ipad 13", "Could mean iPad Pro 13\" or iPad Air 13\"."),
        ]
        html.append("<h2>iPad Query → Reply map (paste this when seller types the matching query)</h2>")
        for query, why in ipad_query_aliases:
            reply = ('To get you a price, which iPad model and storage (e.g., iPad Pro 11" M4 256GB '
                     'or iPad Air 11" M2 128GB), and is it WiFi or Cellular?')
            html.append(f"<p><strong>Query \"{query}\"</strong> ({why}) → Reply: \"{reply}\"</p>")
        html.append("")

    # Per-category quick answers at the top — SHORT and dense so it fits in one chunk.
    if category == "iphone-used":
        html.append("<h2>Naked-query defaults (model only, no storage/lock/condition)</h2>")
        html.append("<p><strong>Rule:</strong> When seller types just a model name (e.g. \"16 pro\", \"17 pro max\", \"15\"), quote the matching line below. Smallest storage, SIM Unlocked, Grade A. End every USED quote with \"Prices can change at any time at our discretion.\"</p>")
        by_model_local = defaultdict(list)
        model_order_local = []
        for entry in entries:
            if entry.condition != "Grade A" or entry.lock != "Unlocked":
                continue
            if entry.model not in by_model_local:
                model_order_local.append(entry.model)
            by_model_local[entry.model].append(entry)
        # Build a map of model -> (best_storage, price) for the query->quote section.
        model_defaults = {}
        for model in model_order_local:
            entries_m = sorted(by_model_local[model], key=lambda e: _storage_kb(e.storage))
            if not entries_m:
                continue
            best = entries_m[0]
            model_defaults[model] = (best.storage, best.price)
            line = (
                f"<p><strong>{model}</strong>: default <strong>${best.price}</strong> "
                f"({best.storage} SIM Unlocked Grade A).</p>"
            )
            html.append(line)
        html.append("")

        # Explicit query->quote map. GPT-4.1 follows literal "when X, reply Y"
        # instructions much more reliably than abstract rules about defaults.
        html.append("<h2>Query → Reply map (paste this answer verbatim when seller types the matching query)</h2>")
        html.append("<p>Each line below is a complete pre-written answer for a short query. If the seller's message is just one of these queries (model name with no extra detail), reply with the exact sentence shown. Do NOT ask 'which exact model?' — even if the input is ambiguous like '17', use the base-model entry.</p>")

        # Generate canonical query phrasings for each model
        def query_aliases(model):
            """Return a list of likely seller phrasings for this model."""
            aliases = []
            m = re.match(r"^iPhone (\d+)\s*(.*)$", model)
            if m:
                num = m.group(1)
                suffix = m.group(2).strip()
                # short numeric form first (most ambiguous, most important)
                if suffix == "":
                    aliases.append(num)
                    aliases.append(f"iphone {num}")
                    aliases.append(f"iPhone {num}")
                elif suffix.lower() == "pro max":
                    aliases.append(f"{num} pro max")
                    aliases.append(f"{num} promax")
                    aliases.append(f"{num} pm")
                    aliases.append(f"{num}pm")
                    aliases.append(f"{num} max")
                    aliases.append(f"iphone {num} pro max")
                    aliases.append(f"iphone {num} pm")
                    aliases.append(f"iPhone {num} Pro Max")
                elif suffix.lower() == "pro":
                    aliases.append(f"{num} pro")
                    aliases.append(f"iphone {num} pro")
                    aliases.append(f"iPhone {num} Pro")
                elif suffix.lower() == "plus":
                    aliases.append(f"{num} plus")
                    aliases.append(f"{num}+")
                    aliases.append(f"{num} +")
                    aliases.append(f"iphone {num} plus")
                    aliases.append(f"iphone {num}+")
                    aliases.append(f"iPhone {num} Plus")
                elif suffix.upper() == "AIR":
                    aliases.append(f"{num} air")
                    aliases.append(f"iphone {num} air")
                    aliases.append(f"iPhone {num} Air")
                elif suffix.upper() == "E":
                    aliases.append(f"{num}e")
                    aliases.append(f"iphone {num}e")
                    aliases.append(f"iPhone {num}E")
                elif suffix.lower() == "mini":
                    aliases.append(f"{num} mini")
                    aliases.append(f"iphone {num} mini")
                    aliases.append(f"iPhone {num} Mini")
            elif "SE" in model:
                aliases.append("se")
                aliases.append("iphone se")
                aliases.append("iPhone SE")
            return aliases

        for model in model_order_local:
            if model not in model_defaults:
                continue
            storage, price = model_defaults[model]
            aliases = query_aliases(model)
            if not aliases:
                continue
            alias_str = " / ".join(f'"{a}"' for a in aliases)
            quote = (
                f"If it's SIM unlocked with no scratches at all, we can offer up to "
                f"${price} for the {storage} {model}. Prices can change at any time at our discretion."
            )
            html.append(
                f"<p><strong>Query {alias_str}</strong> → Reply: \"{quote}\"</p>"
            )
        html.append("")
        html.append("")

    if category == "iphone-used":
        # Group entries by MODEL and emit one section per model.
        # This is critical for GHL retrieval: per-model chunks let multi-model
        # queries pull complete pricing for each model in one chunk, instead of
        # slicing one model across multiple chunks (which the 3-chunk ceiling
        # cannot accommodate).
        by_model = defaultdict(list)
        for entry in entries:
            by_model[entry.model].append(entry)

        # Order: 17 series first (descending generation), then by suffix priority
        def _model_key(model):
            mm = re.match(r"^iPhone (\d+)\s*(.*)$", model)
            if not mm:
                return (0, 99, model)
            n = int(mm.group(1))
            suffix = mm.group(2).strip().lower()
            suffix_order = {"pro max": 0, "pro": 1, "air": 2, "plus": 3, "": 4, "e": 5, "mini": 6}
            return (-n, suffix_order.get(suffix, 99), model)

        for model in sorted(by_model.keys(), key=_model_key):
            section_html = render_per_model_section(model, by_model[model])
            html.append(section_html)
            html.append("")
    else:
        for variant_key in variant_order:
            section_html = render_section(variant_key, by_variant[variant_key])
            html.append(section_html)
            html.append("")

    html.append("<h2>Grading Reference</h2>")
    html.append("<p><strong>Grade A:</strong> Used, no scratches, fully functional. Default for sellers who don't specify condition.</p>")
    html.append("<p><strong>Grade B:</strong> Used, light wear, no cracks, fully functional.</p>")
    html.append("<p><strong>Grade C:</strong> Used, cracked screen / hairline crack / heavy scratches, fully functional.</p>")
    html.append("<p><strong>Grade D:</strong> Used, bad LCD / dead pixels / lines on screen / black spots.</p>")
    html.append("<p><strong>DOA:</strong> Dead — won't power, water damage.</p>")
    html.append("<p><strong>SWAP HSO:</strong> Factory-fresh, zero cycles, but no original box. Never quote unless seller explicitly says \"no box\" + \"never used\".</p>")
    html.append("<p><strong>Sealed:</strong> Factory-sealed in original box, never opened, never activated.</p>")
    html.append("<p><strong>Open Box:</strong> Brand new but box has been opened, device never used.</p>")
    html.append("<p><strong>Sealed (Activated):</strong> Sealed in box but activation has begun.</p>")
    html.append("")
    html.append("</body></html>")

    return "\n".join(html)


def render_aggregate(category_htmls, all_entries):
    """Aggregate prices.html — Quick Answers FIRST, then per-category content."""
    today_utc = datetime.datetime.now(datetime.timezone.utc).isoformat() + "Z"

    html = []
    html.append("<!DOCTYPE html>")
    html.append('<html lang="en"><head><meta charset="UTF-8">')
    html.append("<title>BLT Trading — Mobile Device Price Sheet</title></head>")
    html.append("<body>")
    html.append("<h1>BLT Trading — Mobile Device Price Sheet (Aggregate)</h1>")
    html.append(f"<p><strong>Last Updated:</strong> {today_utc}</p>")
    html.append("")

    # Quick Answers FIRST — most important for naked queries
    html.append(render_quick_answers(all_entries))
    html.append("")
    html.append("<hr/>")
    html.append("<h1>Variant detail sections (per category)</h1>")
    html.append("<p>The sections below are for refining a quote AFTER the seller specifies storage, lock, or condition. For naked model queries with no detail, see Quick Answers above instead.</p>")
    html.append("")

    # Then per-category content (variants, grading reference) — but skip the
    # iphone-used canonical-defaults sub-block since Quick Answers already
    # covers it more visibly at the top.
    for cat, content in category_htmls.items():
        m = re.search(r"<body>(.*)</body>", content, re.DOTALL)
        if not m:
            continue
        body = m.group(1).strip()
        # Strip the iphone-used internal Default Buying Prices header+block
        # (Quick Answers above is already a more visible duplicate).
        if cat == "iphone-used":
            body = re.sub(
                r"<h2>Default Buying Prices.*?(?=<section>)",
                "",
                body,
                count=1,
                flags=re.DOTALL,
            )
        html.append(f"<!-- ===== {cat}.html ===== -->")
        html.append(body)
        html.append("")

    html.append("</body></html>")
    return "\n".join(html)




def render_iphone_defaults(category_entries):
    """Render iphone-defaults.html — short Quick Reference page for retrieval.

    GHL caps retrieval at ~3 chunks per query. With one big iphone-used.html,
    multi-model queries (4+ phones) can lose chunks. This page is small enough
    to fit in ONE chunk and lists every iPhone model's Grade A default in one
    place — so even when GHL drops variant-detail chunks, defaults survive.
    """
    today_utc = datetime.datetime.now(datetime.timezone.utc).isoformat() + "Z"

    models_order = [
        "iPhone 17 Pro Max", "iPhone 17 Pro", "iPhone 17 AIR", "iPhone 17", "iPhone 17E",
        "iPhone 16 Pro Max", "iPhone 16 Pro", "iPhone 16 Plus", "iPhone 16", "iPhone 16E",
        "iPhone 15 Pro Max", "iPhone 15 Pro", "iPhone 15 Plus", "iPhone 15",
        "iPhone 14 Pro Max", "iPhone 14 Pro", "iPhone 14 Plus", "iPhone 14",
        "iPhone 13 Pro Max", "iPhone 13 Pro", "iPhone 13", "iPhone 13 Mini",
        "iPhone SE (3rd Gen)",
        "iPhone 12 Pro Max", "iPhone 12 Pro", "iPhone 12", "iPhone 12 Mini",
        "iPhone 11 Pro Max", "iPhone 11 Pro", "iPhone 11",
    ]

    used = category_entries.get("iphone-used", []) or []

    def storage_to_int(s):
        if not s: return 999999
        s = s.upper().replace(" ", "")
        m2 = re.match(r"(\d+)(GB|TB)?", s)
        if not m2: return 999999
        n = int(m2.group(1))
        return n * 1024 if (m2.group(2) == "TB") else n

    # Index entries by exact canonical model name
    by_model = {canon: [] for canon in models_order}
    for e in used:
        for canon in sorted(models_order, key=lambda x: -len(x)):
            if e.model == canon:
                by_model[canon].append(e)
                break

    def find_price(entries, lock_value, condition):
        matches = [e for e in entries if e.lock == lock_value and e.condition == condition]
        if not matches: return None, None
        m3 = min(matches, key=lambda x: storage_to_int(x.storage))
        return m3.price, m3.storage

    html = ['<!DOCTYPE html><html lang="en"><head><meta charset="UTF-8">',
            '<title>BLT Trading - iPhone Default Prices</title></head><body>',
            '<h1>BLT iPhone Default Buying Prices</h1>',
            f'<p>Updated: {today_utc}. Defaults = Grade A (no scratches), smallest storage. Use this for any iPhone query, especially multi-model messages where the seller names 2+ phones at once.</p>',
            '<h2>iPhone Grade A defaults (every model in one place)</h2>',
            '<ul>']

    for canon in models_order:
        entries = by_model.get(canon, [])
        if not entries: continue
        u_price, u_storage = find_price(entries, "Unlocked", "Grade A")
        l_price, l_storage = find_price(entries, "Carrier Locked", "Grade A")
        if u_price is None and l_price is None: continue
        storage = u_storage or l_storage
        bits = []
        if u_price is not None: bits.append(f"Unlocked ${u_price}")
        if l_price is not None: bits.append(f"Carrier-Locked ${l_price}")
        html.append(f"<li>{canon} {storage}: {', '.join(bits)}.</li>")

    html.append('</ul>')
    html.append('<h2>iPhone NEW Sealed and Open Box defaults (Carrier-Locked, smallest storage)</h2>')
    html.append('<ul>')
    for canon in models_order:
        entries = by_model.get(canon, [])
        s_price, s_storage = find_price(entries, "Carrier Locked", "Sealed")
        o_price, o_storage = find_price(entries, "Carrier Locked", "Open Box")
        if s_price is None and o_price is None: continue
        storage = s_storage or o_storage
        bits = []
        if s_price is not None: bits.append(f"Sealed ${s_price}")
        if o_price is not None: bits.append(f"Open Box ${o_price}")
        html.append(f"<li>{canon} {storage}: {', '.join(bits)}.</li>")
    html.append('</ul>')
    html.append('<p>For storage tiers other than the default, condition grades (B/C/D/DOA), or HSO pricing - see iphone-used.html for the full per-variant breakdown.</p>')
    html.append('</body></html>')
    return '\n'.join(html)


def render_welcome_html():
    """Render welcome/policies page."""
    today_utc = datetime.datetime.now(datetime.timezone.utc).isoformat() + "Z"

    html = []
    html.append("<!DOCTYPE html>")
    html.append('<html lang="en"><head><meta charset="UTF-8">')
    html.append("<title>BLT Trading — Welcome, Policies &amp; Grading Guide</title></head>")
    html.append("<body>")
    html.append("<h1>BLT Trading — Welcome, Policies &amp; Grading Guide</h1>")
    html.append(f"<p><strong>Last Updated:</strong> {today_utc}</p>")
    html.append("<p>BLT Trading buys used and brand-new gadgets — phones, iPads, Apple Watches, AirPods, gaming consoles. We do NOT sell, trade-in, repair, unlock, or activate.</p>")
    html.append("")
    html.append("<h2>Contact Information</h2>")
    html.append("<p><strong>Yu</strong> — call or text (909) 664-5589 (default escalation)</p>")
    html.append("<p><strong>Angelina</strong> — call or text (909) 631-1132 (default escalation)</p>")
    html.append("<p><strong>Nick</strong> — text only (628) 266-5678 (weekends and after-hours only)</p>")
    html.append("<p><strong>Email:</strong> info@BLTtradings.com</p>")
    html.append("<p><strong>Hours:</strong> Mon–Fri 11AM–6PM CST. Closed every Thursday. Saturdays/Sundays by appointment.</p>")
    html.append("<p><strong>Shipping address:</strong> 2955 Congressman Ln, Dallas, TX 75220</p>")
    html.append("")
    html.append("<h2>Payment Methods</h2>")
    html.append("<p>Cash payment available in DFW Metropolitan area only. Outside DFW: wire transfer, ACH, or Zelle.</p>")
    html.append("<p>Apple Gift Cards purchased at 78% of face value.</p>")
    html.append("<p>Free FedEx shipping labels for sellers shipping 5+ devices.</p>")
    html.append("")
    html.append("<h2>General Policy</h2>")
    html.append("<p>Prices can change at any time at our discretion.</p>")
    html.append("<p>All quoted prices are buying prices — what BLT pays the seller.</p>")
    html.append("<p>Final price is confirmed only after in-person inspection or device test.</p>")
    html.append("<p>BLT does not sell devices to customers. BLT does not perform repairs, unlocks, or activations.</p>")
    html.append("")
    html.append("<h2>Specialist Routing</h2>")
    html.append("<p>Models not in the price list: text or call Yu (909) 664-5589 or Angelina (909) 631-1132 to confirm if BLT can still take them.</p>")
    html.append("<p>Bulk inquiries (5+ devices): same — Yu or Angelina.</p>")
    html.append("<p>Weekend or after-hours appointments: text Nick at (628) 266-5678 only. Do not contact Yu or Angelina outside business hours.</p>")
    html.append("")
    html.append("<h2>Device Grading &amp; Term Guide</h2>")
    html.append("<p><strong>NEW Sealed:</strong> Factory-sealed, brand new in box, never opened, never activated.</p>")
    html.append("<p><strong>NEW Open Box:</strong> Brand new but the seal is opened. Never used.</p>")
    html.append("<p><strong>Sealed (Activated):</strong> Sealed in box but Apple activation has occurred (clock started).</p>")
    html.append("<p><strong>SWAP / HSO:</strong> Factory-fresh device, zero battery cycles, but without the original box.</p>")
    html.append("<p><strong>Grade A:</strong> Used, no scratches at all, fully functional.</p>")
    html.append("<p><strong>Grade B:</strong> Used, light wear, no cracks, fully functional.</p>")
    html.append("<p><strong>Grade C:</strong> Used, visible scratches or hairline crack/chip, fully functional.</p>")
    html.append("<p><strong>Grade D:</strong> Used, cracked screen, bad LCD, dead pixels, or heavy damage but powers on.</p>")
    html.append("<p><strong>DOA:</strong> Dead on arrival — won't power on or has water damage.</p>")
    html.append("")
    html.append("</body></html>")
    return "\n".join(html)


# ===== Main =====

def main():
    print("=" * 70)
    print("BLT Trading Price Scraper — Variant-First HTML Generator")
    print("=" * 70)

    if not Path(BLT_XLSX_PATH).exists():
        download_xlsx()
    else:
        print(f"Using cached {BLT_XLSX_PATH}")

    print(f"Loading {BLT_XLSX_PATH}...")
    wb = load_workbook(BLT_XLSX_PATH, data_only=True)

    all_entries = []

    category_data = {
        "iphone-used": ("Used iphone ", parse_iphone_used),
        "iphone-new": ("New Iphone", parse_iphone_new),
        "ipad": (None, lambda ws: parse_ipad_used(ws) + parse_ipad_new(ws)),
        "samsung": ("Samsung", parse_samsung),
        "watch": ("Apple watch ", parse_watch),
        "gaming": ("switch- PS5", parse_gaming),
    }

    category_entries = {}

    for cat, (sheet_name, parser) in category_data.items():
        if sheet_name:
            ws = None
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            else:
                for sn in wb.sheetnames:
                    if sheet_name.lower().strip() in sn.lower().strip():
                        ws = wb[sn]
                        break

            if not ws:
                print(f"  WARN: {sheet_name} not found in workbook")
                category_entries[cat] = []
                continue

            entries = parser(ws)
            category_entries[cat] = entries
            all_entries.extend(entries)
            print(f"  {cat}: {len(entries)} entries")
        else:
            entries = []
            for sn in wb.sheetnames:
                if "ipad" in sn.lower():
                    entries.extend(parser(wb[sn]))
            category_entries[cat] = entries
            all_entries.extend(entries)
            print(f"  {cat}: {len(entries)} entries")

    # MERGE iphone-new entries INTO iphone-used so each variant section has
    # all conditions (Sealed, Open Box, Grade A DEFAULT, B, C, D, DOA, SWAP HSO)
    # in one chunk. Prevents the bot from picking NEW Sealed for naked queries
    # because Grade A is right there as the explicit DEFAULT.
    new_iphone_entries = category_entries.get("iphone-new", [])
    used_iphone_entries = category_entries.get("iphone-used", [])
    # Re-tag NEW entries so they appear inside the USED file's sections.
    for e in new_iphone_entries:
        e.category = "iphone-used"
    category_entries["iphone-used"] = used_iphone_entries + new_iphone_entries
    # Empty iphone-new so its file becomes a stub (do not index this URL anymore).
    category_entries["iphone-new"] = []

    print(f"\nMerged {len(new_iphone_entries)} NEW iPhone entries into iphone-used sections")
    print(f"Total entries parsed: {len(all_entries)}")

    category_htmls = {}

    category_info = {
        "iphone-new": ("BLT Trading — iPhone Buying Prices (Factory-Fresh)", "iphone-new"),
        "iphone-used": ("BLT Trading — iPhone Buying Prices (USED Grades)", "iphone-used"),
        "ipad": ("BLT Trading — iPad Buying Prices", "ipad"),
        "samsung": ("BLT Trading — Samsung Buying Prices", "samsung"),
        "watch": ("BLT Trading — Apple Watch Buying Prices", "watch"),
        "gaming": ("BLT Trading — Gaming Console Buying Prices", "gaming"),
    }

    print("\nGenerating HTML files...")

    for cat, (title, filename) in category_info.items():
        entries = category_entries.get(cat, [])
        content = render_category_html(cat, title, entries)
        category_htmls[cat] = content

        outpath = OUTDIR / f"{filename}.html"
        outpath.write_text(content)
        size_kb = len(content) / 1024
        section_count = content.count("<section>")
        print(f"  {filename}.html: {len(entries)} entries, {section_count} sections, {size_kb:.1f}KB")

    welcome_content = render_welcome_html()
    outpath = OUTDIR / "welcome.html"
    outpath.write_text(welcome_content)
    print(f"  welcome.html: {len(welcome_content) / 1024:.1f}KB")

    iphone_defaults_content = render_iphone_defaults(category_entries)
    outpath = OUTDIR / "iphone-defaults.html"
    outpath.write_text(iphone_defaults_content)
    print(f"  iphone-defaults.html: {len(iphone_defaults_content) / 1024:.1f}KB")

    print("\n" + "=" * 70)
    print(f"SUCCESS: Generated 8 HTML files in {OUTDIR}")
    print("=" * 70)


if __name__ == "__main__":
    main()
